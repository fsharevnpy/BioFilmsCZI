# ===== crash logging early =====
import os, sys, traceback, tempfile, datetime, faulthandler, atexit


def resource_path(rel):
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, rel)


# Try ICO (Windows), fall back to PNG (cross-platform)
ico = resource_path("BioAnalyzer.ico")
png = resource_path("BioAnalyzer.png")

LOG_PATH = os.path.join(
    tempfile.gettempdir(),
    f"biofilms_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
)

# Keep the file handle alive until process exit
_FH = open(LOG_PATH, "w", buffering=1)
try:
    faulthandler.enable(_FH)  # native faults go here
except Exception:
    pass


def _log_exception():
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write("\n--- exception ---\n")
            traceback.print_exc(file=f)
    except Exception:
        pass


def _excepthook(exc_type, exc, tb):
    _log_exception()
    try:
        import tkinter as _tk
        from tkinter import messagebox as _mb

        _tk.Tk().withdraw()
        _mb.showerror("Error", f"An error occurred.\nLog: {LOG_PATH}")
    except Exception:
        pass
    # still propagate for console builds
    sys.__excepthook__(exc_type, exc, tb)


sys.excepthook = _excepthook


@atexit.register
def _close_log():
    try:
        _FH.close()
    except Exception:
        pass

# ===== end crash logging =====

import math
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Tuple

import cv2  # ensure cv2 is available for all functions below
import numpy as np
from PIL import Image, ImageTk

# Enable OpenCV optimizations
try:
    cv2.setUseOptimized(True)
except Exception:
    pass

try:
    # Avoid oversubscription when we also use ThreadPoolExecutor
    cv2.setNumThreads(0)
except Exception:
    pass

# ---------- Analysis helpers (scalars & color compose) ----------

def normalize_8bit(image: np.ndarray, lower: float = 1, upper: float = 99) -> np.uint8:
    """Percentile-based 8-bit normalization with epsilon guard."""
    p_low, p_high = np.percentile(image, (lower, upper))
    scale = (p_high - p_low) + 1e-6
    img_scaled = np.clip((image - p_low) / scale * 255, 0, 255)
    return img_scaled.astype(np.uint8)


def compose_rgb(
    dead_brightest: np.ndarray,
    live_brightest: np.ndarray,
    dead_count: int,
    live_count: int,
) -> np.ndarray:
    """
    Scale channels by counts and compose BGR image (G = live, R = dead).
    NOTE: Returns BGR array to be consistent with OpenCV color order.
    """
    eps = 1e-9
    ratio = (dead_count + eps) / (live_count + eps)
    clip_ratio = np.clip(ratio, 1 / math.pi, math.pi)

    dead_scale = math.sqrt(clip_ratio)
    live_scale = 1.0 / math.sqrt(clip_ratio)

    red_raw = dead_brightest.astype(np.float32) * dead_scale
    green_raw = live_brightest.astype(np.float32) * live_scale

    bgr = np.zeros((*dead_brightest.shape, 3), dtype=np.uint8)
    bgr[..., 1] = normalize_8bit(green_raw)  # G: live
    bgr[..., 2] = normalize_8bit(red_raw)  # R: dead
    return bgr


def intensity_rgb(image: np.ndarray) -> Tuple[float, float, float]:
    """Measure red/green pixel coverage (%) in a BGR image; return (combined%, red%, green%)."""
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Red has two ranges in HSV
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 70, 50])
    upper_red2 = np.array([180, 255, 255])

    # Green range
    lower_green = np.array([40, 70, 50])
    upper_green = np.array([80, 255, 255])

    mask_red1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask_red2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask_red = cv2.bitwise_or(mask_red1, mask_red2)
    mask_green = cv2.inRange(hsv, lower_green, upper_green)

    red_pixels = int(np.sum(mask_red > 0))
    green_pixels = int(np.sum(mask_green > 0))
    total_pixels = int(image.shape[0] * image.shape[1]) if image.size else 1

    red_percentage = (red_pixels / total_pixels) * 100.0
    green_percentage = (green_pixels / total_pixels) * 100.0
    combined_percentage = red_percentage + green_percentage
    return combined_percentage, red_percentage, green_percentage

# ---------- Fast-path caches & reusable kernels ----------
KERNEL_3X3_RECT = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
KERNEL_3X3_CROSS = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
ONES_3X3 = np.ones((3, 3), dtype=np.uint8)

USE_FAST_DENOISE = True  # True: GaussianBlur + Otsu (fast). False: fastNlMeansDenoising (slower)
PREFER_XIMGPROC = True  # Use cv2.ximgproc.thinning (Guo-Hall) if available

# ---- Globals (app state) ----
file_list: List[str] = []

# Old per-session image context retained for compatibility with some helpers
czi_obj = None
image_array = None
axes_labels = ""

# GUI state
base_pil = None
zoom = 1.0
min_zoom = 0.05
max_zoom = 10.0
offset_x = 0
offset_y = 0
drag_start = None
last_wm_state = None
current_view = "image"  # "image" or "info"
meta_dict = {}

# Simple info vars (basic summary shown in Info view too)
file_var = None
shape_var = None
axes_var = None
cnum_var = None
znum_var = None

_current_c_idx = 0
_current_z_idx = 0

# ===== Precomputed store for ALL selected files =====
# Structure:
# precomputed_store[file_key] = {
#   "path": str, "axes": str, "shape": tuple,
#   "meta": dict,
#   "nC": int, "nZ": int,
#   "planes": {(c,z): PIL.Image},              # raw planes
#   "overlay": {(c,z): PIL.Image},             # overlay images (Mark)
#   "overlay_stats": {(c,z): dict or None},    # overlay stats
#   "merge": {z: PIL.Image},                   # merged image R/G
#   "merge_stats": {z: dict},                  # merged stats
# }
precomputed_store: Dict[str, Dict[str, Any]] = {}
current_file_key: str = ""


# ---------- Axes helpers & plane extraction ----------

def infer_axes(shape, info):
    # Try to read dimension order from info; fallback to heuristic
    if isinstance(info, dict):
        ax = info.get("axes") or info.get("dimension_order") or info.get("dims")
        if isinstance(ax, str) and len(ax) == len(shape):
            return ax
    n = len(shape)
    if n == 2:
        return "YX"
    if n == 3:
        return "CYX" if any(s in (3, 4) for s in shape) else "ZYX"
    if n == 4:
        return "CZYX"
    if n == 5:
        return "TCZYX"
    if n == 6:
        return "STCZYX"
    if n == 7:
        return "STRCZYX"
    return "YX".rjust(n, "S")


def get_count(axis, shape, axes):
    return int(shape[axes.index(axis)]) if axis in axes else 0


def normalize_to_uint8(arr: np.ndarray) -> np.ndarray:
    # Fast min-max normalization to uint8 (fallback to zeros if invalid)
    if arr.dtype == np.uint8:
        return arr
    a = arr.astype(np.float32, copy=False)
    vmin = float(np.nanmin(a))
    vmax = float(np.nanmax(a))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax <= vmin:
        return np.zeros_like(a, dtype=np.uint8)
    a = (a - vmin) * (255.0 / (vmax - vmin))
    return a.astype(np.uint8)


def extract_plane(arr, axes, c_idx=0, z_idx=0):
    # Extract a single 2D plane with optional channel and z selection
    if "Y" not in axes or "X" not in axes:
        raise ValueError(f"Missing Y/X in axes: {axes}")
    idx = [slice(None)] * len(axes)
    for ax in ("S", "T", "R", "M"):
        if ax in axes:
            idx[axes.index(ax)] = 0
    if "C" in axes:
        nC = int(arr.shape[axes.index("C")])
        if nC > 0:
            c_idx = max(0, min(int(c_idx), nC - 1))
            idx[axes.index("C")] = c_idx
    if "Z" in axes:
        nZ = int(arr.shape[axes.index("Z")])
        if nZ > 0:
            z_idx = max(0, min(int(z_idx), nZ - 1))
            idx[axes.index("Z")] = z_idx
    plane = np.squeeze(arr[tuple(idx)])
    if plane.ndim == 2:
        return normalize_to_uint8(plane)
    if plane.ndim == 3 and plane.shape[-1] in (3, 4):
        if plane.dtype != np.uint8:
            plane = np.stack([normalize_to_uint8(plane[..., k]) for k in range(plane.shape[-1])], axis=-1)
        return plane[..., :3]
    return normalize_to_uint8(np.squeeze(plane))


def to_uint8_slice(image_array: np.ndarray, ch: int, z: int, axes: str) -> np.uint8:
    """Extract a 2D plane by dynamic axes and convert to uint8."""
    plane = extract_plane(image_array, axes, c_idx=ch, z_idx=z)
    return plane if plane.dtype == np.uint8 else normalize_to_uint8(plane)

# ---------- Binary mask & contours ----------

def denoise_and_threshold(image: np.ndarray) -> np.uint8:
    """
    Fast denoise + threshold for bright cells.
    USE_FAST_DENOISE=True -> GaussianBlur + Otsu (fast)
    else -> fastNlMeansDenoising (slower, slightly smoother)
    """
    if USE_FAST_DENOISE:
        blur = cv2.GaussianBlur(image, (3, 3), 0, borderType=cv2.BORDER_REPLICATE)
        _, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return th
    else:
        den = cv2.fastNlMeansDenoising(image, None, h=15, templateWindowSize=5, searchWindowSize=15)
        _, th = cv2.threshold(den, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return th


def find_external_contours(binary: np.ndarray) -> List[np.ndarray]:
    """Find external contours from a binary mask."""
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    return contours


def is_closed_contour(contour: np.ndarray, threshold: float = 10.0) -> bool:
    """Check if a contour is near-closed by distance between first/last points."""
    start_point = contour[0][0]
    end_point = contour[-1][0]
    distance = float(np.linalg.norm(start_point - end_point))
    return distance < threshold


def keep_max_pixel_per_contour(image: np.ndarray, contours: List[np.ndarray], close_thresh: float = 10.0) -> np.uint8:
    """For each closed contour: keep only pixels equal to the max intensity inside that contour."""
    out = np.zeros_like(image)
    for c in contours:
        if is_closed_contour(c, close_thresh):
            c_mask = np.zeros_like(image)
            cv2.drawContours(c_mask, [c], -1, 255, thickness=cv2.FILLED)
            inside = image[c_mask > 0]
            if inside.size > 0:
                m = int(np.max(inside))
                out[(c_mask > 0) & (image == m)] = m
    return out

# ---------- Skeletonization & branch points ----------

def _morph_skeleton_8c(bin_img: np.ndarray) -> np.uint8:
    """Morphological skeletonization using 8-connectivity (fast loop)."""
    skel = np.zeros_like(bin_img, dtype=np.uint8)
    img = bin_img.copy()
    while True:
        eroded = cv2.erode(img, KERNEL_3X3_RECT)
        opened = cv2.dilate(eroded, KERNEL_3X3_RECT)
        temp = cv2.subtract(img, opened)
        skel |= temp
        img = eroded
        if cv2.countNonZero(img) == 0:
            break
    return skel


# ---------- ROI skeleton ----------

def skeletonize_inside_closed_contours(
    image: np.ndarray,
    contours: List[np.ndarray],
    close_thresh: float = 10.0,
    prefer_ximgproc: bool = PREFER_XIMGPROC,
) -> Tuple[np.uint8, int]:
    skel_full = np.zeros(image.shape, dtype=np.uint8)
    use_ximgproc = prefer_ximgproc and hasattr(cv2, "ximgproc") and hasattr(cv2.ximgproc, "thinning")
    num_total = 0
    for c in contours:
        if not is_closed_contour(c, close_thresh):
            continue
        x, y, w, h = cv2.boundingRect(c)
        roi = image[y : y + h, x : x + w]
        c_mask = np.zeros((h, w), dtype=np.uint8)
        cv2.drawContours(c_mask, [c - [x, y]], -1, 255, thickness=cv2.FILLED)
        inside = cv2.bitwise_and(roi, roi, mask=c_mask)
        bin_roi = (inside > 0).astype(np.uint8) * 255
        if use_ximgproc:
            skel_roi = cv2.ximgproc.thinning(bin_roi, thinningType=cv2.ximgproc.THINNING_GUOHALL)
        else:
            skel_roi = _morph_skeleton_8c(bin_roi)
        skel_roi = ((skel_roi > 0).astype(np.uint8)) * 255
        skel_full[y : y + h, x : x + w] |= skel_roi
        num_labels, _ = cv2.connectedComponents((skel_roi > 0).astype(np.uint8), connectivity=8)
        num_total += max(0, num_labels - 1)
    return skel_full, num_total


def detect_branch_points(skeleton_u8: np.uint8) -> np.uint8:
    sk = (skeleton_u8 > 0).astype(np.uint8)
    neighbor_sum = cv2.filter2D(sk, ddepth=cv2.CV_8U, kernel=ONES_3X3, borderType=cv2.BORDER_CONSTANT)
    neighbor_cnt = neighbor_sum - sk
    return ((sk == 1) & (neighbor_cnt > 2)).astype(np.uint8) * 255


def overlay_skeleton(origin_gray: np.uint8, skeleton_u8: np.uint8, branch_points_u8: np.uint8 = None) -> np.ndarray:
    rgb = cv2.cvtColor(origin_gray, cv2.COLOR_GRAY2BGR)
    rgb[skeleton_u8 > 0] = [0, 100, 255]
    if branch_points_u8 is not None:
        rgb[branch_points_u8 > 0] = [255, 0, 0]
    return rgb

# ---------- Precompute for a file (independent of global per-session caches) ----------
from aicspylibczi import CziFile

from concurrent.futures import ThreadPoolExecutor, as_completed
import os

def precompute_one_file(path: Path, on_progress=None, max_workers: int = None) -> Dict[str, Any]:
    local: Dict[str, Any] = {
        "path": str(path),
        "axes": "",
        "shape": (),
        "meta": {},
        "nC": 0,
        "nZ": 0,
        "planes": {},
        "overlay": {},
        "overlay_stats": {},
        "merge": {},
        "merge_stats": {},
    }

    czi = CziFile(str(path))
    try:
        arr, info = czi.read_image()
        axes = infer_axes(arr.shape, info)
        nC = get_count("C", arr.shape, axes)
        nZ = get_count("Z", arr.shape, axes)
        local["axes"]  = axes
        local["shape"] = tuple(arr.shape)
        local["nC"]    = nC
        local["nZ"]    = nZ
        try:
            local["meta"] = parse_czi_metadata_from_czi(czi)
        except Exception:
            local["meta"] = {}
    finally:
        try:
            czi.close()
        except Exception:
            pass

    cs = range(nC) if ("C" in axes and nC > 0) else [0]
    zs = range(nZ) if ("Z" in axes and nZ > 0) else [0]
    needed_cs = [0, 1] if nC >= 2 else [0]

    total_steps = max(1, len(needed_cs) * len(zs))
    step = 0
    if on_progress:
        on_progress(step, total_steps)

    if max_workers is None:
        try:
            max_workers = max(1, min(4, (os.cpu_count() or 2) - 1))
        except Exception:
            max_workers = 2

    def _proc_cz(_c: int, _z: int) -> Dict[str, Any]:
        origin_u8 = to_uint8_slice(arr, _c, _z, axes)
        th = denoise_and_threshold(origin_u8)
        contours = find_external_contours(th)
        brightest = keep_max_pixel_per_contour(origin_u8, contours, 10.0)
        skeleton_u8, num_skel = skeletonize_inside_closed_contours(brightest, contours, 10.0)
        branch_u8 = detect_branch_points(skeleton_u8)
        overlay_bgr = overlay_skeleton(origin_u8, skeleton_u8, branch_u8)
        return {
            "origin_u8": origin_u8,
            "brightest_u8": brightest,
            "overlay_bgr": overlay_bgr,
            "num_skeletons": int(num_skel),
        }

    for z in zs:
        for c in cs:
            plane = extract_plane(arr, axes, c_idx=c, z_idx=z)
            local["planes"][(int(c), int(z))] = Image.fromarray(plane)

    for z in zs:
        proc_cache: Dict[int, Dict[str, Any]] = {}
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futs = {ex.submit(_proc_cz, c, z): c for c in needed_cs}
            for fut in as_completed(futs):
                c = futs[fut]
                out = fut.result()
                proc_cache[c] = out

                overlay_rgb = cv2.cvtColor(out["overlay_bgr"], cv2.COLOR_BGR2RGB)
                local["overlay"][(c, z)] = Image.fromarray(overlay_rgb)

                which = "dead" if c == 0 else ("live" if c == 1 else "other")
                if which in ("dead", "live"):
                    cur_count = int(out["num_skeletons"])
                    dead_count = cur_count if which == "dead" else 0
                    live_count = cur_count if which == "live" else 0
                    tot = dead_count + live_count
                    dead_prop = (dead_count / tot) if tot > 0 else 0.0
                    live_prop = (live_count / tot) if tot > 0 else 0.0

                    cur_u8 = out["brightest_u8"]
                    zero = np.zeros_like(cur_u8)
                    bgr_image = compose_rgb(
                        dead_brightest=cur_u8 if which == "dead" else zero,
                        live_brightest=cur_u8 if which == "live" else zero,
                        dead_count=dead_count,
                        live_count=live_count,
                    )
                    percentage, _, _ = intensity_rgb(bgr_image)

                    local["overlay_stats"][(c, z)] = {
                        "which": which,
                        "dead_count": dead_count,
                        "live_count": live_count,
                        "dead_prop": dead_prop,
                        "live_prop": live_prop,
                        "percentage": percentage,
                    }
                else:
                    local["overlay_stats"][(c, z)] = None

                step += 1
                if on_progress:
                    on_progress(step, total_steps)

        if (0 in proc_cache) and (1 in proc_cache):
            dead = proc_cache[0]
            live = proc_cache[1]
            rgb_bgr = compose_rgb(
                dead_brightest=dead["brightest_u8"],
                live_brightest=live["brightest_u8"],
                dead_count=dead["num_skeletons"],
                live_count=live["num_skeletons"],
            )
            local["merge"][int(z)] = Image.fromarray(cv2.cvtColor(rgb_bgr, cv2.COLOR_BGR2RGB))

            dead_count = int(dead["num_skeletons"])
            live_count = int(live["num_skeletons"])
            tot = dead_count + live_count
            if tot > 0:
                dead_prop = dead_count / tot
                live_prop = live_count / tot
                ratio = (dead_count / live_count) if live_count > 0 else float("inf")
            else:
                dead_prop = live_prop = 0.0
                ratio = None
            percentage, _, _ = intensity_rgb(rgb_bgr)
            local["merge_stats"][int(z)] = {
                "dead_count": dead_count,
                "live_count": live_count,
                "dead_prop": dead_prop,
                "live_prop": live_prop,
                "ratio": ratio,
                "percentage": percentage,
            }
        else:
            local["merge"][int(z)] = None
            local["merge_stats"][int(z)] = None

    return local

# ---------- Metadata parsing ----------
# --------- Export to Excel ---------
def _sanitize_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    cleaned = ''.join((ch if ch not in bad else '_') for ch in name)
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned or 'Sheet'

def set_row_height_pixels(ws, row_idx: int, px: int):
    ws.row_dimensions[row_idx].height = max(1, px * 0.75)

def set_col_width_pixels(ws, col_idx: int, px: int):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = max(1, (px + 5) / 7)

def export_to_excel():
    if not precomputed_store:
        messagebox.showwarning("Export", "No CZI files to export")
        return

    ok = messagebox.askyesno("Export", "Are you sure to export all CZI files?")
    if not ok:
        return

    try:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

        used_names = set()
        for key, entry in precomputed_store.items():
            base = Path(entry.get("path", key)).stem
            name = _sanitize_sheet_name(base)
            orig = name; k = 2
            while name in used_names:
                suff = f"_{k}"
                name = _sanitize_sheet_name(orig[: (31 - len(suff))] + suff)
                k += 1
            used_names.add(name)

            ws = wb.create_sheet(title=name)
            # header
            ws.append([
                "Merged Image",
                "Live (count)",
                "Dead (count)",
                "Live (%)",
                "Dead (%)",
                "Intensity (%)"
            ])

            header_row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=header_row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font = Font(bold=True)

            nZ = int(entry.get("nZ", 0) or 0)
            mstats = entry.get("merge_stats") or {}
            merge_imgs = entry.get("merge") or {}

            max_img_w = 0
            for z in range(nZ):
                st = mstats.get(z)
                row = z + 2

                if st:
                    live_ct = int(st.get("live_count", 0) or 0)
                    dead_ct = int(st.get("dead_count", 0) or 0)
                    live_pc = round(float(st.get("live_prop", 0.0) * 100.0), 2)
                    dead_pc = round(float(st.get("dead_prop", 0.0) * 100.0), 2)
                    intensity = round(float(st.get("percentage", 0.0) or 0.0), 6)

                    ws.cell(row=row, column=2, value=live_ct).alignment = Alignment(horizontal="center", vertical="top")
                    ws.cell(row=row, column=3, value=dead_ct).alignment = Alignment(horizontal="center", vertical="top")
                    ws.cell(row=row, column=4, value=live_pc).alignment = Alignment(horizontal="center", vertical="top")
                    ws.cell(row=row, column=5, value=dead_pc).alignment = Alignment(horizontal="center", vertical="top")
                    ws.cell(row=row, column=6, value=intensity).alignment = Alignment(horizontal="center", vertical="top")

                pil_img = merge_imgs.get(z)  # PIL.Image
                if pil_img:
                    bio = io.BytesIO()
                    pil_img.save(bio, format="PNG")
                    bio.seek(0)

                    xl_img = XLImage(bio)
                    xl_img.width = pil_img.width
                    xl_img.height = pil_img.height

                    xl_img.anchor = f"A{row}"
                    ws.add_image(xl_img)

                    set_row_height_pixels(ws, row, pil_img.height)

                    if pil_img.width > max_img_w:
                        max_img_w = pil_img.width
                else:
                    set_row_height_pixels(ws, row, 24)

            if max_img_w > 0:
                set_col_width_pixels(ws, 1, max_img_w)

            for col in range(2, 7):
                set_col_width_pixels(ws, col, 110)

        out_path = os.path.join(
            os.getcwd(),
            f"bio_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(out_path)
        messagebox.showinfo("Export", f"Exportation Completed :\n{out_path}")

    except Exception as e:
        _log_exception()
        messagebox.showerror("Export", f"Error during exportation:\n{e}\n\nLog: {LOG_PATH}")

def parse_czi_metadata_from_czi(czi: CziFile) -> dict:
    """Extract a robust subset of CZI metadata (works with Element or ElementTree)."""
    root = czi.meta
    try:
        _find = root.find
        _findall = root.findall
    except AttributeError:
        root = root.getroot()
        _find = root.find
        _findall = root.findall

    def find_text(elem, path):
        if elem is None:
            return None
        tag = elem.find(path)
        return tag.text.strip() if tag is not None and tag.text else None

    result = {}
    image = _find("Metadata/Information/Image")
    result["SizeX"] = find_text(image, "SizeX")
    result["SizeY"] = find_text(image, "SizeY")
    result["SizeZ"] = find_text(image, "SizeZ")
    result["SizeC"] = find_text(image, "SizeC")
    result["PixelType"] = find_text(image, "PixelType")

    result["PixelSize"] = {}
    for dist in _findall("Metadata/Scaling/Items/Distance"):
        axis = dist.attrib.get("Id") if dist is not None else None
        val = find_text(dist, "Value")
        if axis and val:
            try:
                result["PixelSize"][axis] = float(val)
            except ValueError:
                result["PixelSize"][axis] = val

    result["Channels"] = []
    for ch in _findall("Metadata/Information/Image/Dimensions/Channels/Channel"):
        det_node = ch.find("DetectionWavelength")
        detection = None
        if det_node is not None:
            detection = find_text(det_node, "Ranges")
        result["Channels"].append(
            {
                "Name": ch.attrib.get("Name"),
                "Fluor": find_text(ch, "Fluor"),
                "Excitation": find_text(ch, "ExcitationWavelength"),
                "Emission": find_text(ch, "EmissionWavelength"),
                "DetectionRange": detection,
            }
        )

    obj = _find("Metadata/Information/Instrument/Objectives/Objective")
    manuf = obj.find("Manufacturer") if obj is not None else None
    result["Objective"] = (
        {
            "Model": find_text(manuf, "Model") if manuf is not None else None,
            "NA": find_text(obj, "LensNA") if obj is not None else None,
            "Magnification": find_text(obj, "NominalMagnification") if obj is not None else None,
            "Immersion": find_text(obj, "Immersion") if obj is not None else None,
        }
        if obj is not None
        else {}
    )

    scope = _find("Metadata/Information/Instrument/Microscopes/Microscope")
    result["Microscope"] = find_text(scope, "System") if scope is not None else None
    return result


# ---------- GUI & interactions ----------
import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    import io
except Exception as _e:
    Workbook = None

def _file_key_from_path(p: Path) -> str:
    try:
        return str(p.resolve())
    except Exception:
        return str(p)


def _set_header_vars_from_store(entry: dict):
    file_var.set(Path(entry["path"]).name)
    shape_var.set(str(tuple(entry["shape"])))
    axes_var.set(entry["axes"])
    cnum_var.set(str(entry["nC"]))
    znum_var.set(str(entry["nZ"]))


# ---- View (cover fit, clamp, zoom & pan) ----

def fit_cover_view_to_canvas():
    # Auto fit the image to cover the canvas area while preserving aspect ratio
    global zoom, min_zoom, offset_x, offset_y
    if base_pil is None:
        return
    iw, ih = base_pil.size
    cw = max(1, canvas.winfo_width() or 800)
    ch = max(1, canvas.winfo_height() or 420)
    if iw == 0 or ih == 0:
        return
    cover = max(cw / iw, ch / ih)
    min_zoom = cover
    zoom = max(zoom, cover)
    sw, sh = int(iw * zoom), int(ih * zoom)
    offset_x = (cw - sw) // 2
    offset_y = (ch - sh) // 2


def clamp_offsets():
    # Prevent panning beyond the image edges
    global offset_x, offset_y
    if base_pil is None:
        return
    iw, ih = base_pil.size
    sw = max(1, int(iw * zoom))
    sh = max(1, int(ih * zoom))
    cw = max(1, canvas.winfo_width() or 800)
    ch = max(1, canvas.winfo_height() or 420)
    offset_x = min(0, max(cw - sw, offset_x))
    offset_y = min(0, max(ch - sh, offset_y))


def redraw_canvas():
    if _IN_PRECOMPUTE:
        return
    # Draw the current PIL image onto the canvas
    global display_imgtk
    canvas.delete("all")
    if base_pil is None:
        draw_placeholder()
        return
    clamp_offsets()
    iw, ih = base_pil.size
    sw = max(1, int(iw * zoom))
    sh = max(1, int(ih * zoom))
    disp = base_pil.resize((sw, sh), Image.Resampling.LANCZOS)
    display_imgtk = ImageTk.PhotoImage(disp)
    canvas.create_image(offset_x, offset_y, image=display_imgtk, anchor="nw")

    # Draw legend after image is drawn
    draw_legend_on_canvas()


def canvas_reset_view(event=None):
    # Reset zoom and recenter
    if current_view != "image":
        return
    fit_cover_view_to_canvas()
    redraw_canvas()


def canvas_start_drag(event):
    # Begin panning
    global drag_start
    if current_view != "image":
        return
    drag_start = (event.x, event.y)


def canvas_drag(event):
    # Continue panning
    global offset_x, offset_y, drag_start
    if current_view != "image":
        return
    if drag_start is None or base_pil is None:
        return
    dx = event.x - drag_start[0]
    dy = event.y - drag_start[1]
    offset_x += dx
    offset_y += dy
    drag_start = (event.x, event.y)
    redraw_canvas()


def canvas_end_drag(event):
    # End panning
    global drag_start
    drag_start = None


def canvas_zoom(event, factor=None):
    # Zoom with mouse wheel or programmatic factor
    global zoom, offset_x, offset_y
    if current_view != "image" or base_pil is None:
        return
    if factor is None:
        if getattr(event, "num", None) == 4:
            factor = 1.1
        elif getattr(event, "num", None) == 5:
            factor = 1 / 1.1
        else:
            factor = 1.1 if getattr(event, "delta", 0) > 0 else 1 / 1.1
    new_zoom = max(min(zoom * factor, max_zoom), min_zoom)
    if abs(new_zoom - zoom) < 1e-6:
        return
    mx, my = event.x, event.y
    img_x = (mx - offset_x) / zoom
    img_y = (my - offset_y) / zoom
    zoom = new_zoom
    offset_x = mx - img_x * zoom
    offset_y = my - img_y * zoom
    redraw_canvas()

# ---------- Placeholder ----------

def draw_placeholder():
    # Draw a neutral placeholder when no image is loaded
    cw = max(1, canvas.winfo_width() or 800)
    ch = max(1, canvas.winfo_height() or 420)
    pad = 12
    side = min(cw, ch) - 2 * pad
    side = max(side, 60)
    x0 = (cw - side) // 2
    y0 = (ch - side) // 2
    x1 = x0 + side
    y1 = y0 + side
    canvas.create_rectangle(x0, y0, x1, y1, outline="#444", width=2, fill="#1b1b1b")
    canvas.create_text((cw // 2), (ch // 2) - 12, text="No image loaded", fill="#bbb", font=("Helvetica", 14, "bold"))
    canvas.create_text((cw // 2), (ch // 2) + 12, text="Click 'Browse files' to open .czi", fill="#888", font=("Helvetica", 11))

# ---------- Canvas sizing to image aspect ----------

def size_canvas_to_aspect(aspect: float):
    # Adjust canvas widget size to match image aspect ratio
    root.update_idletasks()
    left_w = left.winfo_width() or 0
    win_w = max(1, root.winfo_width())
    win_h = max(1, root.winfo_height())
    avail_w = max(1, win_w - left_w)
    avail_h = max(1, win_h)
    if avail_w / avail_h > aspect:
        new_h = avail_h
        new_w = int(new_h * aspect)
    else:
        new_w = avail_w
        new_h = int(new_w / aspect)
    canvas.config(width=new_w, height=new_h)
    canvas_container.config(width=new_w, height=new_h)


def refresh_layout_and_view():
    # Recompute layout and redraw the image
    global zoom
    if _IN_PRECOMPUTE:
        return
    if current_view != "image":
        return
    if base_pil is None:
        size_canvas_to_aspect(1.0)
        redraw_canvas()
        return
    iw, ih = base_pil.size
    aspect = iw / ih if ih != 0 else 1.0
    size_canvas_to_aspect(aspect)
    prev_zoom = zoom
    fit_cover_view_to_canvas()
    zoom = prev_zoom if prev_zoom > min_zoom else min_zoom
    redraw_canvas()

# ---------- Info view rendering ----------

def render_info_view():
    # Render a scrollable metadata view
    for w in info_inner.winfo_children():
        w.destroy()

    tk.Label(
        info_inner, text="Image Information", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 14, "bold")
    ).pack(anchor="w", pady=(0, 8))

    def kv_row(parent, key, val):
        rowf = tk.Frame(parent, bg="#1a1a1a")
        rowf.pack(fill="x", pady=1)
        tk.Label(rowf, text=f"{key}", fg="#bbb", bg="#1a1a1a", width=14, anchor="w").pack(side="left")
        tk.Label(rowf, text=val, fg="#eee", bg="#1a1a1a", anchor="w", wraplength=900, justify="left").pack(
            side="left", fill="x", expand=True
        )

    imf = tk.Frame(info_inner, bg="#1a1a1a")
    imf.pack(fill="x", pady=(0, 4))
    tk.Label(imf, text="Image:", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 12, "bold")).pack(anchor="w")

    def add_if(k, v):
        if v is not None and v != "":
            kv_row(imf, f"  {k}", str(v))

    add_if("SizeX", meta_dict.get("SizeX"))
    add_if("SizeY", meta_dict.get("SizeY"))
    add_if("SizeZ", meta_dict.get("SizeZ"))
    add_if("SizeC", meta_dict.get("SizeC"))
    add_if("PixelType", meta_dict.get("PixelType"))

    ps = meta_dict.get("PixelSize") or {}
    if ps:
        psf = tk.Frame(info_inner, bg="#1a1a1a")
        psf.pack(fill="x", pady=(4, 4))
        tk.Label(psf, text="PixelSize:", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 12, "bold")).pack(
            anchor="w"
        )
        for axis, val in ps.items():
            kv_row(psf, f"  {axis}", str(val))

    chs = meta_dict.get("Channels") or []
    chf = tk.Frame(info_inner, bg="#1a1a1a")
    chf.pack(fill="x", pady=(4, 4))
    tk.Label(chf, text=f"Channels ({len(chs)}):", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 12, "bold")).pack(
        anchor="w"
    )
    if chs:
        for i, ch in enumerate(chs, 1):
            blk = tk.Frame(chf, bg="#1a1a1a", highlightthickness=0)
            blk.pack(fill="x", pady=(2, 2))
            kv_row(blk, f"  [{i}] Name", str(ch.get("Name") or ""))

            def chrow(k, v):
                kv_row(blk, f"    {k}", str(v) if v is not None else "")

            chrow("Fluor", ch.get("Fluor"))
            chrow("Excitation", ch.get("Excitation"))
            chrow("Emission", ch.get("Emission"))
            chrow("DetectionRange", ch.get("DetectionRange"))
    else:
        kv_row(chf, "  (none)", "")

    obj = meta_dict.get("Objective") or {}
    of = tk.Frame(info_inner, bg="#1a1a1a")
    of.pack(fill="x", pady=(4, 4))
    tk.Label(of, text="Objective:", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 12, "bold")).pack(anchor="w")
    if obj:
        for k in ["Model", "NA", "Magnification", "Immersion"]:
            kv_row(of, f"  {k}", str(obj.get(k)) if obj.get(k) is not None else "")
    else:
        kv_row(of, "  (none)", "")

    ms = meta_dict.get("Microscope")
    mf = tk.Frame(info_inner, bg="#1a1a1a")
    mf.pack(fill="x", pady=(4, 0))
    tk.Label(mf, text="Microscope:", fg="#ddd", bg="#1a1a1a", font=("Helvetica", 12, "bold")).pack(anchor="w")
    kv_row(mf, "  System", str(ms) if ms is not None else "")

    info_inner.update_idletasks()
    info_canvas.configure(scrollregion=info_canvas.bbox("all"))

# ---------- Legend drawer (now uses precomputed_store) ----------

def draw_legend_on_canvas():
    if base_pil is None or current_view != "image":
        return
    if not current_file_key or current_file_key not in precomputed_store:
        return

    entry = precomputed_store[current_file_key]
    lines = None

    if tick_merge.get():
        st = entry["merge_stats"].get(int(_current_z_idx))
        if not st:
            return
        ratio = st["ratio"]
        if ratio is None:
            ratio_str = "N/A"
        elif ratio == float("inf"):
            ratio_str = "∞"
        else:
            ratio_str = f"{ratio:.3f}"
        lines = [
            f"Dead: {st['dead_prop']:.2%} ({st['dead_count']})",
            f"Live: {st['live_prop']:.2%} ({st['live_count']})",
            f"Dead/Live: {ratio_str}",
            f"Intensity: {st['percentage']:.1f}%",
        ]

    elif tick_analyze.get():
        st = entry["overlay_stats"].get((int(_current_c_idx), int(_current_z_idx)))
        if not st:
            return
        if st.get("which") == "dead":
            lines = [f"Dead: {st['dead_count']} cell(s)"]
        elif st.get("which") == "live":
            lines = [f"Live: {st['live_count']} cell(s)"]
        else:
            return
    else:
        return

    text = "\n".join(lines)
    pad_x = 10
    pad_y = 30
    x = (canvas.winfo_width() or 0) - pad_x
    y = (canvas.winfo_height() or 0) - pad_y

    text_id = canvas.create_text(
        x,
        y,
        text=text,
        fill="#eeeeee",
        font=("Helvetica", 10, "bold"),
        anchor="se",
        tags=("legend_text",),
        justify="right",
    )
    bbox = canvas.bbox(text_id)
    if bbox:
        x0, y0, x1, y1 = bbox
        bg_pad = 6
        rect_id = canvas.create_rectangle(
            x0 - bg_pad,
            y0 - bg_pad,
            x1 + bg_pad,
            y1 + bg_pad,
            fill="#000000",
            outline="#666666",
            width=1,
            tags=("legend_bg",),
        )
        canvas.tag_lower(rect_id, text_id)


# ---------- Preview update (only uses precomputed_store) ----------

def update_preview():
    global base_pil, zoom, _current_c_idx, _current_z_idx, meta_dict
    if not current_file_key or current_file_key not in precomputed_store:
        if current_view == "image":
            redraw_canvas()
        return

    entry = precomputed_store[current_file_key]
    try:
        c_idx = int(spin_c.get()) if (spin_c["state"] == "normal") else 0
        z_idx = int(spin_z.get()) if (spin_z["state"] == "normal") else 0
    except Exception:
        c_idx, z_idx = 0, 0

    _current_c_idx, _current_z_idx = c_idx, z_idx

    # 1) Merge  2) Mark  3) Raw
    if tick_merge.get():
        img = entry["merge"].get(z_idx)
        if img is None:
            img = entry["planes"].get((c_idx, z_idx))
        base_pil = img
    elif tick_analyze.get():
        img = entry["overlay"].get((c_idx, z_idx))
        if img is None:
            img = entry["planes"].get((c_idx, z_idx))
        base_pil = img
    else:
        base_pil = entry["planes"].get((c_idx, z_idx))

    # Keep meta in sync for Info view
    meta_dict = entry["meta"]

    zoom = 1.0
    if current_view == "image":
        refresh_layout_and_view()


# ---------- Multi-file + GUI plumbing ----------

def on_select_file(event=None):
    global current_file_key, meta_dict
    sel = listbox.curselection()
    if not sel:
        return
    idx = sel[0]
    if idx >= len(file_list):
        return
    key = _file_key_from_path(Path(file_list[idx]))
    if key not in precomputed_store:
        messagebox.showerror("Error", f"No cached data for {key}.")
        return

    current_file_key = key
    entry = precomputed_store[key]
    _set_header_vars_from_store(entry)

    spin_c.config(
        from_=0,
        to=max(0, entry["nC"] - 1),
        state="normal" if (entry["nC"] > 1 and not tick_merge.get()) else "disabled",
        wrap=True,
        increment=1,
    )
    spin_c.delete(0, "end")
    spin_c.insert(0, "0")
    spin_z.config(
        from_=0,
        to=max(0, entry["nZ"] - 1),
        state="normal" if entry["nZ"] > 1 else "disabled",
        wrap=True,
        increment=1,
    )
    spin_z.delete(0, "end")
    spin_z.insert(0, "0")

    _apply_c_visibility()
    meta_dict = entry["meta"]
    update_preview()

import threading, queue, time

# === Progress bus (UI-safe) ===
class _ProgressBus:
    def __init__(self):
        self.q = queue.Queue()
        self.done = threading.Event()
        self.err = None
        self.current_file_hdr = None  # (i, total_files, fname, total_steps)
        self.last_step = 0
        self.last_total = 1

    def post_header(self, i, nfiles, fname, total_steps):
        self.current_file_hdr = (i, nfiles, fname, max(1, int(total_steps)))
        self.q.put(("hdr", self.current_file_hdr))

    def post_step(self, step, total):
        self.last_step, self.last_total = int(step), max(1, int(total))
        self.q.put(("step", (self.last_step, self.last_total)))

    def post_done(self):
        self.q.put(("done", None))
        self.done.set()

    def post_error(self, exc):
        self.err = exc
        self.q.put(("error", str(exc)))
        self.done.set()

PROGRESS_POLL_MS = 33  # ~30 FPS

class FileSlicesProgress:
    def __init__(self, master: tk.Tk, bus: _ProgressBus):
        self.bus = bus
        self.top = tk.Toplevel(master)
        self.top.title("Loading ...")
        self.top.transient(master)
        self.top.grab_set()
        self.top.resizable(False, False)
        try:
            self.top.attributes("-topmost", True)
        except Exception:
            pass
        self.cancelled = False

        frm = tk.Frame(self.top, padx=12, pady=12)
        frm.pack(fill="both", expand=True)

        self.var_hdr = tk.StringVar(value="Starting…")
        tk.Label(frm, textvariable=self.var_hdr, font=("Helvetica", 10, "bold")).pack(anchor="w", pady=(0, 6))

        self.pb = ttk.Progressbar(frm, mode="determinate", length=380, maximum=1, value=0)
        self.pb.pack(fill="x")

        btns = tk.Frame(frm); btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="Cancel", command=self._cancel).pack(side="right")

        self.top.protocol("WM_DELETE_WINDOW", self._cancel)
        self._poll_id = None
        self._start_poll()

    def _start_poll(self):
        try:
            while True:
                kind, payload = self.bus.q.get_nowait()
                if kind == "hdr":
                    i, nf, fname, tot = payload
                    self.var_hdr.set(f"File {i}/{nf}: {fname}")
                    self.pb.configure(maximum=tot, value=0)
                elif kind == "step":
                    step, tot = payload
                    self.pb.configure(maximum=tot, value=min(step, tot))
                elif kind == "error":
                    self.var_hdr.set(f"Error: {payload}")
                elif kind == "done":
                    pass
        except queue.Empty:
            pass

        if not self.bus.done.is_set() and not self.cancelled:
            self._poll_id = self.top.after(PROGRESS_POLL_MS, self._start_poll)
        else:
            try:
                self.pb.configure(value=min(self.bus.last_step, self.bus.last_total),
                                  maximum=self.bus.last_total)
            except Exception:
                pass

    def begin_file(self, file_idx: int, files_total: int, fname: str, total_steps: int):
        self.bus.post_header(file_idx, files_total, fname, total_steps)

    def update_slice(self, step: int, total_steps: int):
        self.bus.post_step(step, total_steps)

    def _cancel(self):
        self.cancelled = True

    def close(self):
        if self._poll_id:
            try:
                self.top.after_cancel(self._poll_id)
            except Exception:
                pass
        try:
            self.top.grab_release()
        except Exception:
            pass
        self.top.destroy()

# ===== UI performance helpers (no algorithm changes) =====
UI_THROTTLE_MS = 50
_IN_PRECOMPUTE = False
_last_ui_tick_ms = 0

def ui_set_busy(is_busy: bool):
    cur = "watch" if is_busy else ""
    try:
        root.config(cursor=cur)
        canvas.config(cursor="watch" if is_busy else "tcross")
    except Exception:
        pass

    state = "disabled" if is_busy else "normal"
    try:
        btn_browse.config(state=state)
        btn_image.config(state=state)
        btn_info.config(state=state)
        btn_export.config(state=state if Workbook is not None else "disabled")
        chk1.config(state=state)  # Mark
        chk2.config(state=state)  # Merge
        spin_c.config(state="disabled" if is_busy else spin_c.cget("state"))
        spin_z.config(state="disabled" if is_busy else spin_z.cget("state"))
        listbox.config(state=state)
    except Exception:
        pass

    global _IN_PRECOMPUTE
    _IN_PRECOMPUTE = is_busy

def ui_throttled_progress(pb_widget, value, maximum):
    import time
    global _last_ui_tick_ms
    now = int(time.time() * 1000)
    if now - _last_ui_tick_ms < UI_THROTTLE_MS and value < maximum:
        return
    _last_ui_tick_ms = now
    try:
        pb_widget.configure(value=min(value, maximum))
        pb_widget.update_idletasks()
    except Exception:
        pass
# ===== end UI helpers =====

# Replace browse_files to precompute ALL selected files synchronously

def browse_files():
    global file_list, current_file_key, meta_dict
    try:
        fns = filedialog.askopenfilenames(
            title="Choose CZI files",
            filetypes=[("Zeiss CZI files", "*.czi"), ("All files", "*.*")]
        )
    except Exception:
        _log_exception()
        messagebox.showerror("Browse error", f"Failed to open file dialog.\nSee log:\n{LOG_PATH}")
        return
    if not fns:
        return

    precomputed_store.clear()
    file_list = list(fns)
    listbox.delete(0, tk.END)
    for f in file_list:
        listbox.insert(tk.END, Path(f).name)

    # === UI busy state
    ui_set_busy(True)

    # === Progress bus & dialog
    bus = _ProgressBus()
    dlg = FileSlicesProgress(root, bus)
    processed_any = False

    def _worker():
        try:
            for i, f in enumerate(file_list, start=1):
                p = Path(f)
                fname = p.name

                try:
                    _tmp = CziFile(str(p))
                    _arr, _info = _tmp.read_image()
                    _axes = infer_axes(_arr.shape, _info)
                    _nC = get_count("C", _arr.shape, _axes)
                    _nZ = get_count("Z", _arr.shape, _axes)
                    _needed_cs = [0, 1] if _nC >= 2 else [0]
                    total = max(1, len(_needed_cs) * (_nZ if _nZ > 0 else 1))
                    try:
                        _tmp.close()
                    except Exception:
                        pass
                except Exception:
                    total = 1

                bus.post_header(i, len(file_list), fname, total)
                if dlg.cancelled:
                    break

                def _on_progress(step, tot):
                    bus.post_step(step, tot)

                try:
                    entry = precompute_one_file(p, on_progress=_on_progress)
                    precomputed_store[_file_key_from_path(p)] = entry
                    nonlocal processed_any
                    processed_any = True
                except Exception as e:
                    _log_exception()
                    bus.post_error(e)
                    continue

                if dlg.cancelled:
                    break
        except Exception as e:
            bus.post_error(e)
        finally:
            bus.post_done()

    # === Start worker
    t = threading.Thread(target=_worker, daemon=True)
    t.start()

    def _finish_when_done():
        if not bus.done.is_set() and not dlg.cancelled:
            root.after(50, _finish_when_done)
            return

        try:
            dlg.close()
        except Exception:
            pass
        ui_set_busy(False)

        if not processed_any:
            return

        listbox.selection_clear(0, tk.END)
        listbox.selection_set(0)
        key0 = _file_key_from_path(Path(file_list[0]))
        entry = precomputed_store[key0]
        globals()["current_file_key"] = key0
        _set_header_vars_from_store(entry)

        spin_c.config(
            from_=0, to=max(0, entry["nC"] - 1),
            state="normal" if (entry["nC"] > 1 and not tick_merge.get()) else "disabled",
            wrap=True, increment=1
        )
        spin_c.delete(0, "end"); spin_c.insert(0, "0")

        spin_z.config(
            from_=0, to=max(0, entry["nZ"] - 1),
            state="normal" if entry["nZ"] > 1 else "disabled",
            wrap=True, increment=1
        )
        spin_z.delete(0, "end"); spin_z.insert(0, "0")

        _apply_c_visibility()
        globals()["meta_dict"] = entry["meta"]
        update_preview()

    _finish_when_done()

def _debug_open_czi(path: Path):
    try:
        czi = CziFile(str(path))
    except Exception:
        _log_exception()
        messagebox.showerror("CZI error (open)", f"Failed at CziFile(...)\n{path}\n\nLog: {LOG_PATH}")
        return

    try:
        _ = czi.meta
    except Exception:
        _log_exception()
        messagebox.showerror("CZI error (meta)", f"CziFile.meta failed\n{path}\n\nLog: {LOG_PATH}")
        return

    try:
        # No longer using load_czi; precompute_one_file handles reading
        pass
    except Exception:
        _log_exception()
        return


# ---------- Toggle between Image and Info ----------

def set_view(view_name: str):
    global current_view, zoom
    if view_name == current_view:
        return
    current_view = view_name
    if current_view == "image":
        btn_image.config(relief="sunken")
        btn_info.config(relief="raised")
        info_frame.pack_forget()
        canvas_stack.pack(fill="both", expand=True)
        canvas_container.pack(fill="both", expand=True)
        zoom = 1.0
        refresh_layout_and_view()
    else:
        btn_image.config(relief="raised")
        btn_info.config(relief="sunken")
        canvas_container.pack_forget()
        canvas_stack.pack(fill="both", expand=True)
        info_frame.pack(fill="both", expand=True)
        render_info_view()


# ---------- C selector visibility (use store instead of session array) ----------

def _apply_c_visibility():
    if tick_merge.get():
        spin_c.config(state="disabled")
        try:
            lbl_c.config(fg="#777")
        except Exception:
            pass
    else:
        if current_file_key and current_file_key in precomputed_store:
            nC = precomputed_store[current_file_key]["nC"]
        else:
            nC = 0
        spin_c.config(state="normal" if nC > 1 else "disabled")
        try:
            lbl_c.config(fg="black")
        except Exception:
            pass


# ---------- Tick callbacks ----------

def on_tick_analyze():
    # Make analyze & merge mutually exclusive
    if tick_analyze.get() and tick_merge.get():
        tick_merge.set(False)
        _apply_c_visibility()
    update_preview()


def on_tick_merge():
    # Make merge & analyze mutually exclusive
    if tick_merge.get() and tick_analyze.get():
        tick_analyze.set(False)
    _apply_c_visibility()
    update_preview()


# ---------- Build GUI ----------
root = tk.Tk()
root.title("Bio Analyzer")
root.geometry("1100x720")

# ICO first (works if ICO is classic BMP-based and was bundled)
try:
    if os.path.exists(ico):
        root.iconbitmap(ico.replace("\\", "/"))
except Exception:
    pass

# Always set PNG as a fallback (Tk supports PNG via PhotoImage)
try:
    if os.path.exists(png):
        _icon_png = tk.PhotoImage(file=png)
        root.iconphoto(True, _icon_png)
except Exception:
    pass

last_wm_state = root.state()

# Left panel
left = tk.Frame(root, padx=6, pady=6)
left.pack(side="left", fill="y")

controls = tk.Frame(left)
controls.pack(fill="x")
btn_browse = tk.Button(controls, text="Browse files", command=browse_files)
btn_browse.pack(side="left", fill="x", expand=True)

listbox = tk.Listbox(left, height=15, exportselection=False)
listbox.pack(fill="both", expand=True, pady=6)
listbox.bind("<<ListboxSelect>>", on_select_file)

# Container frame right under the listbox
frame_bottom = tk.Frame(left)
frame_bottom.pack(fill="x", pady=(0, 6))

# First checkbox → overlay (mark)
tick_analyze = tk.BooleanVar(value=False)
chk1 = tk.Checkbutton(frame_bottom, text="Mark", variable=tick_analyze, command=on_tick_analyze)
chk1.pack(side="left")

# Second checkbox → merge dead/live and hide C
tick_merge = tk.BooleanVar(value=False)
chk2 = tk.Checkbutton(frame_bottom, text="Merge", variable=tick_merge, command=on_tick_merge)
chk2.pack(side="left", padx=(10, 0))

sel = tk.Frame(left, pady=6)
sel.pack(fill="x")
lbl_c = tk.Label(sel, text="C:")
lbl_c.pack(side="left")
spin_c = tk.Spinbox(sel, from_=0, to=0, width=5, state="disabled", wrap=True, increment=1, command=update_preview)
spin_c.pack(side="left")
tk.Label(sel, text="Z:").pack(side="left", padx=(8, 0))
spin_z = tk.Spinbox(sel, from_=0, to=0, width=5, state="disabled", wrap=True, increment=1, command=update_preview)
spin_z.pack(side="left")

# Right side (toolbar + stack area)
right = tk.Frame(root)
right.pack(side="left", fill="both", expand=True)

toolbar = tk.Frame(right, pady=4)
toolbar.pack(fill="x")
btn_image = tk.Button(toolbar, text="Image", width=8, relief="sunken", command=lambda: set_view("image"))
btn_info = tk.Button(toolbar, text="Info", width=8, relief="raised", command=lambda: set_view("info"))
btn_export = tk.Button(toolbar, text="Export", width=8, command=export_to_excel)

btn_image.pack(side="left", padx=(0, 4))
btn_info.pack(side="left")
btn_export.pack(side="left", padx=(4, 0))

# Stack area
canvas_stack = tk.Frame(right)
canvas_stack.pack(fill="both", expand=True)

# --- Image view ---
canvas_container = tk.Frame(canvas_stack, bg="#111")
canvas_container.pack(fill="both", expand=True)
canvas = tk.Canvas(canvas_container, bg="#222", cursor="tcross", highlightthickness=0)
canvas.place(relx=0.5, rely=0.5, anchor="center")

# --- Info view (scrollable) ---
info_frame = tk.Frame(canvas_stack, bg="#1a1a1a")
info_canvas = tk.Canvas(info_frame, bg="#1a1a1a", highlightthickness=0)
info_scrollbar = tk.Scrollbar(info_frame, orient="vertical", command=info_canvas.yview)
info_canvas.configure(yscrollcommand=info_scrollbar.set)
info_scrollbar.pack(side="right", fill="y")
info_canvas.pack(side="left", fill="both", expand=True)
info_inner = tk.Frame(info_canvas, bg="#1a1a1a")
info_window = info_canvas.create_window((0, 0), window=info_inner, anchor="nw")


def _on_info_configure(event):
    # Keep inner frame width equal to visible width for better wrapping
    info_canvas.itemconfig(info_window, width=info_canvas.winfo_width())
    info_canvas.configure(scrollregion=info_canvas.bbox("all"))


info_canvas.bind("<Configure>", _on_info_configure)

# Mouse wheel bindings for INFO view (scroll)

def info_on_mousewheel(event):
    if current_view != "info":
        return
    if getattr(event, "num", None) == 4:
        info_canvas.yview_scroll(-3, "units")
    elif getattr(event, "num", None) == 5:
        info_canvas.yview_scroll(3, "units")
    else:
        direction = -1 if event.delta > 0 else 1
        info_canvas.yview_scroll(direction * 3, "units")


# Bind wheel events to info_canvas
info_canvas.bind("<MouseWheel>", info_on_mousewheel)  # Windows/Mac
info_canvas.bind("<Button-4>", info_on_mousewheel)  # Linux up
info_canvas.bind("<Button-5>", info_on_mousewheel)  # Linux down

# Simple info vars init
file_var = tk.StringVar(value="-")
shape_var = tk.StringVar(value="-")
axes_var = tk.StringVar(value="-")
cnum_var = tk.StringVar(value="-")
znum_var = tk.StringVar(value="-")

# Bindings (image view)
canvas.bind("<MouseWheel>", canvas_zoom)
canvas.bind("<Button-4>", canvas_zoom)
canvas.bind("<Button-5>", canvas_zoom)
canvas.bind("<Button-1>", canvas_start_drag)
canvas.bind("<B1-Motion>", canvas_drag)
canvas.bind("<ButtonRelease-1>", canvas_end_drag)
canvas.bind("<Double-Button-1>", canvas_reset_view)

spin_c.bind("<Return>", lambda e: update_preview())
spin_c.bind("<FocusOut>", lambda e: update_preview())
spin_z.bind("<Return>", lambda e: update_preview())
spin_z.bind("<FocusOut>", lambda e: update_preview())


def kb_zoom_in(event):
    if current_view != "image":
        return
    class E:
        pass

    e = E()
    e.x = canvas.winfo_width() // 2
    e.y = canvas.winfo_height() // 2
    e.delta = 120
    canvas_zoom(e)


def kb_zoom_out(event):
    if current_view != "image":
        return
    class E:
        pass

    e = E()
    e.x = canvas.winfo_width() // 2
    e.y = canvas.winfo_height() // 2
    e.delta = -120
    canvas_zoom(e)


def kb_reset(event):
    canvas_reset_view()


root.bind("+", kb_zoom_in)
root.bind("-", kb_zoom_out)
root.bind("<space>", kb_reset)

# Maximize/restore → auto cover-fit (image view)

def on_root_resize(event):
    if _IN_PRECOMPUTE:
        return
    global last_wm_state
    cur = root.state()
    if cur != last_wm_state:
        last_wm_state = cur
        canvas_reset_view()
        return
    refresh_layout_and_view()


root.bind("<Configure>", on_root_resize)

# Graceful close

def on_close():
    try:
        pass
    finally:
        root.destroy()


root.protocol("WM_DELETE_WINDOW", on_close)

# Initial placeholder layout/draw
refresh_layout_and_view()
root.mainloop()
